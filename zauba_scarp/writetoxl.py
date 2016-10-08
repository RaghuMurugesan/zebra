#!/usr/bin/env python

import openpyxl
from openpyxl import Workbook
import json

with open('data_gold.json') as data_file:
    data = json.load(data_file)

wb = Workbook()
ws = wb.create_sheet('Gold_imports')
row_index = 1


column_titles = ["COUNTRY OF Discharge", "DESCRIPTION", "PER_UNIT_INR", "VALUE_INR", "DATE", "HS_CODE", "QUANTITY", "UNIT", "PORT_OF_ENTRY"]
for column_index, column_title in zip(range(1, len(column_titles) + 1), column_titles):
    ws.cell(row = row_index, column = column_index).value = column_title
row_index += 1

for line in data:
    #write data
    for cell_data, column_index in zip(line.values(), range(1, line.values().__len__() + 1)):
        #write data_file
        ws.cell(row = row_index, column = column_index).value = cell_data
    row_index += 1
wb.save('test_gold_import.xlsx')
